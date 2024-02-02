import openpyxl

def get_excel_value(file_name: str) -> list:
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active
    list_1 = []
    for sheet_value in ws["A1:C1"]:
        for index, cell in enumerate(sheet_value):
            list_1.append(sheet_value[index].value)   

    return list_1 

def maximaze_matrix(matrix: list) -> list:
    matrix.sort(reverse=True)
    return matrix

def write_to_exel(list_1: list) -> None:
    wb = openpyxl.Workbook()

    ws = wb.active
    for index, rows in enumerate(list_1):
        for col_index, col in enumerate(rows):
            ws[f"{chr(65 + col_index)}{index + 1}"] = col

    wb.save("new.xlsx")

if __name__ == "__main__":

    list_1 = [get_excel_value("1111.xlsx"),get_excel_value("2222.xlsx"), get_excel_value("3333.xlsx")]

    maximaze_matrix(list_1)

    write_to_exel(list_1)
